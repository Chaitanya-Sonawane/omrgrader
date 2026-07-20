"""Excel (color-coded) and PDF report generation."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFEB9C"
GREY = "D9D9D9"

STATUS_FILL = {
    "correct": PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid"),
    "wrong": PatternFill(start_color=RED, end_color=RED, fill_type="solid"),
    "blank": PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid"),
    "multiple": PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid"),
    "not_detected": PatternFill(start_color=GREY, end_color=GREY, fill_type="solid"),
}


def build_excel_report(student_results: list) -> bytes:
    """student_results: list[scoring.StudentResult]"""
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    headers = ["Student ID", "Name", "Correct", "Wrong", "Blank", "Multiple",
               "Not Detected", "Total Marks", "Max Marks", "Percentage"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for r in student_results:
        ws.append([r.student_id, r.student_name, r.correct, r.wrong, r.blank,
                   r.multiple, r.not_detected, r.total_marks, r.max_marks, r.percentage])

    for i, w in enumerate([14, 22, 9, 9, 9, 10, 12, 12, 11, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Per-student detail sheets ---
    for r in student_results:
        safe_name = (r.student_name or r.student_id or "Student")[:25]
        sheet_name = "".join(c for c in safe_name if c.isalnum() or c in " _-")[:28] or "Student"
        # avoid duplicate sheet names
        base_name, n = sheet_name, 1
        while sheet_name in wb.sheetnames:
            n += 1
            sheet_name = f"{base_name[:25]}_{n}"
        ws2 = wb.create_sheet(sheet_name)
        ws2.append(["Q#", "Correct Answer", "Student Answer", "Status", "Marks"])
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for qr in r.question_results:
            row = [qr.question, qr.correct_answer or "-", qr.student_answer or "-",
                   qr.status.replace("_", " ").title(), qr.marks]
            ws2.append(row)
            fill = STATUS_FILL.get(qr.status)
            if fill:
                for c in ws2[ws2.max_row]:
                    c.fill = fill
        for i, w in enumerate([6, 15, 15, 14, 8], start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_report(student_results: list, dashboard: dict | None = None) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("OMR Result Report", styles["Title"]))
    story.append(Spacer(1, 8))

    if dashboard:
        dash_lines = [
            f"Students scanned: {dashboard.get('students_scanned')}",
            f"Highest marks: {dashboard.get('highest_marks')}",
            f"Lowest marks: {dashboard.get('lowest_marks')}",
            f"Average marks: {dashboard.get('average_marks')}",
            f"Pass %: {dashboard.get('pass_percent')}  |  Fail %: {dashboard.get('fail_percent')}",
        ]
        for line in dash_lines:
            story.append(Paragraph(line, styles["Normal"]))
        story.append(Spacer(1, 12))

    table_data = [["Student ID", "Name", "Correct", "Wrong", "Blank", "Total", "Max", "%"]]
    for r in student_results:
        table_data.append([r.student_id, r.student_name, r.correct, r.wrong, r.blank,
                            r.total_marks, r.max_marks, f"{r.percentage}%"])
    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
    ]))
    story.append(t)
    doc.build(story)
    return buf.getvalue()
